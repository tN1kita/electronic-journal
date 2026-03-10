import json
from io import BytesIO

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.import_session import ImportSession
from app.models.student import Student
from app.models.journal_student import JournalStudent
from app.schemas.import_ import ImportMapping


def _cell(row: tuple, col_1based: int):
    i = col_1based - 1
    return row[i] if 0 <= i < len(row) else None


def preview_import(db: Session, journal_id: int, teacher_id: int, file_bytes: bytes, mapping: ImportMapping):
    """
    1) читаем файл
    2) применяем сопоставление
    3) если где-то ошибка — сразу 400
    4) сохраняем ImportSession и возвращаем предпросмотр
    """
    wb = load_workbook(filename=BytesIO(file_bytes))
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Invalid file: no data rows")

    # первая строка — шапка
    data_rows = rows[1:]

    normalized = []
    for r in data_rows:
        num = _cell(r, mapping.number_col)
        surname = _cell(r, mapping.surname_col)
        name = _cell(r, mapping.name_col)
        patronymic = _cell(r, mapping.patronymic_col)
        email = _cell(r, mapping.email_col)
        phone = _cell(r, mapping.phone_col)

        if not surname or not name:
            raise HTTPException(status_code=400, detail="Invalid file: empty surname/name")

        try:
            student_number = int(num) if num is not None else 0
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid file: bad student number")

        surname_s = str(surname).strip()
        name_s = str(name).strip()
        patronymic_s = str(patronymic).strip() if patronymic else ""
        email_s = str(email).strip() if email else None
        phone_s = str(phone).strip() if phone else None

        # минимальная проверка email (строгую можно сделать на фронте)
        if email_s and "@" not in email_s:
            raise HTTPException(status_code=400, detail="Invalid file: bad email")

        fio_short = f"{surname_s} {(name_s[:1] + '.') if name_s else ''}{(patronymic_s[:1] + '.') if patronymic_s else ''}".strip()

        normalized.append(
            {
                "student_number": student_number,
                "surname": surname_s,
                "name": name_s,
                "patronymic": patronymic_s,
                "email": email_s,
                "phone": phone_s,
                "fio_short": fio_short,
            }
        )

    # сохраняем import session
    session = ImportSession(
        journal_id=journal_id,
        teacher_id=teacher_id,
        rows_json=json.dumps(normalized, ensure_ascii=False),
    )
    db.add(session)
    db.commit()
    db.refresh(session)

    preview = normalized[:20]
    return session.id, len(normalized), preview


def confirm_import(db: Session, import_session: ImportSession) -> int:
    """Применяем сохранённый предпросмотр в БД."""
    rows = json.loads(import_session.rows_json)
    imported = 0

    for row in rows:
        email = row.get("email")
        student = None

        if email:
            student = db.query(Student).filter(Student.email == email).first()

        if not student:
            student = Student(
                surname=row["surname"],
                name=row["name"],
                patronymic=row.get("patronymic") or "",
                email=email,
                phone=row.get("phone"),
            )
            db.add(student)
            db.flush()

        # привязка к журналу
        link = (
            db.query(JournalStudent)
            .filter(JournalStudent.journal_id == import_session.journal_id, JournalStudent.student_id == student.id)
            .first()
        )
        if link:
            # обновим номер
            link.student_number = row["student_number"]
        else:
            db.add(
                JournalStudent(
                    journal_id=import_session.journal_id,
                    student_id=student.id,
                    student_number=row["student_number"],
                )
            )
            imported += 1

    db.commit()
    return imported